import re
import os
import sys
import time
import traceback
import tempfile
import tkinter as tk
from tkinter import messagebox, filedialog, scrolledtext
from collections import defaultdict, Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.comments import Comment
import pandas as pd
from tkinterdnd2 import TkinterDnD, DND_FILES  # 拖拽支持

# ==================== 全局样式（单色蓝色，匹配参考图） ====================
BLUE = "#165DFF"
BG_COLOR = "#f5f7fa"
RED = PatternFill("solid", fgColor="FF0000")
YELLOW = PatternFill("solid", fgColor="FFFF00")
RED_CELL = PatternFill("solid", fgColor="FFCCCC")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ==================== 兼容老版.xls的加载函数 ====================
def load_workbook_compat(file_path):
    """自动兼容.xlsx和.xls文件，自动转格式"""
    if file_path.lower().endswith('.xls'):
        # 老版xls文件，自动转成临时xlsx处理
        df = pd.read_excel(file_path)
        # 创建临时文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            temp_path = tmp.name
        df.to_excel(temp_path, index=False)
        # 加载临时xlsx
        wb = load_workbook(temp_path)
        # 用完删临时文件
        os.unlink(temp_path)
        return wb
    else:
        # 正常xlsx，直接加载
        return load_workbook(file_path)

# ==================== 日志输出函数 ====================
def log(msg, log_widget):
    """实时输出日志到界面"""
    log_widget.config(state=tk.NORMAL)
    log_widget.insert(tk.END, f"{msg}\n")
    log_widget.see(tk.END)
    log_widget.update()

# ==================== 解析拖拽文件（处理空格/中文路径） ====================
def parse_drop_files(data):
    """解析拖拽的文件路径，处理带空格的路径"""
    files = []
    i = 0
    n = len(data)
    while i < n:
        if data[i] == '{':
            i += 1
            j = data.find('}', i)
            files.append(data[i:j])
            i = j + 1
        else:
            j = i
            while j < n and data[j] != ' ':
                j += 1
            if i != j:
                files.append(data[i:j])
            i = j + 1
    return files

# ==================== 工具1：发票提取 + 税额核对 ====================
def tool1_extract_and_check(log_widget, file_a=None, file_b=None):
    try:
        log("==================================================", log_widget)
        log("    发票提取 + 税额匹配 二合一工具（财务专用版）", log_widget)
        log("    支持格式：.xlsx / .xls 全兼容", log_widget)
        log("==================================================", log_widget)
        log("", log_widget)

        # 选择文件（如果是拖拽来的就不用选了）
        if not file_a:
            log("📂 请选择【A表 - 待处理表格】", log_widget)
            file_a = filedialog.askopenfilename(
                title="选择A表", 
                filetypes=[("Excel文件", "*.xlsx;*.xls")]
            )
        if not file_a:
            log("❌ 未选择A表，操作取消", log_widget)
            return
        
        if not file_b:
            log("📂 请选择【B表 - 税额对照表格】", log_widget)
            file_b = filedialog.askopenfilename(
                title="选择B表", 
                filetypes=[("Excel文件", "*.xlsx;*.xls")]
            )
        if not file_b:
            log("❌ 未选择B表，操作取消", log_widget)
            return

        log(f"✅ 已选择文件：\n  A表: {file_a}\n  B表: {file_b}", log_widget)
        log("\n⏳ 开始处理：提取发票 + 税额匹配...", log_widget)

        # 兼容加载
        wb = load_workbook_compat(file_a)
        ws = wb.active
        summary_col = None
        header_row = None
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value or "").strip()
                if "摘要" in val:
                    summary_col = cell.column
                    header_row = cell.row
                    log(f"✅ 识别摘要列：第{header_row}行 第{summary_col}列", log_widget)
                    break
            if summary_col:
                break
        if not summary_col:
            log("❌ 未找到【摘要】列", log_widget)
            messagebox.showerror("错误", "未找到【摘要】列")
            return

        left_col = summary_col - 1
        target_col = summary_col + 1
        ws.insert_cols(target_col)
        ws.cell(row=header_row, column=target_col).value = "提取发票号"
        ws.column_dimensions[ws.cell(1, target_col).column_letter].width = 25

        invoice_map = {}
        max_row = ws.max_row
        prev_left = None
        prev_sum = None
        end_row = max_row - 2
        for r in range(header_row + 1, end_row + 1):
            left_val = str(ws.cell(r, left_col).value or "").strip()
            sum_val = str(ws.cell(r, summary_col).value or "").strip()
            if left_val == prev_left and sum_val == prev_sum:
                continue
            prev_left = left_val
            prev_sum = sum_val
            nums = re.findall(r"\d+", sum_val)
            valid = [n for n in nums if len(n) >= 15]
            if valid:
                invoice_map[r] = valid

        for old_row in sorted(invoice_map.keys(), reverse=True):
            invs = invoice_map[old_row]
            if len(invs) == 1:
                ws.cell(old_row, target_col).value = invs[0]
                ws.cell(old_row, target_col).number_format = "@"
            else:
                ws.cell(old_row, target_col).value = invs[-1]
                ws.cell(old_row, target_col).number_format = "@"
                for i in range(len(invs)-2, -1, -1):
                    ir = old_row + 1
                    ws.insert_rows(ir)
                    for c in range(1, summary_col + 1):
                        ws.cell(ir, c).value = ws.cell(old_row, c).value
                    ws.cell(ir, target_col).value = invs[i]
                    ws.cell(ir, target_col).number_format = "@"

        all_inv = []
        max_row = ws.max_row
        end_row = max_row - 2
        for r in range(header_row + 1, end_row + 1):
            v = str(ws.cell(r, target_col).value or "").strip()
            if v:
                all_inv.append(v)
        cnt = Counter(all_inv)
        for r in range(header_row + 1, end_row + 1):
            cell = ws.cell(r, target_col)
            inv = str(cell.value or "").strip()
            if not inv:
                continue
            if len(inv) != 20:
                cell.fill = RED
                cell.comment = Comment("错误发票", "系统")
            if cnt[inv] >= 2:
                cell.fill = YELLOW
                cell.comment = Comment("重复发票", "系统")

        log("🔧 读取B表税额数据...", log_widget)
        # pandas自动兼容xls/xlsx
        df_b = pd.read_excel(file_b, dtype=str)
        b_map = dict(zip(df_b["数电发票号码"], df_b["有效抵扣税额*"]))
        tax_col = 13
        diff_col = 14
        ws.cell(row=header_row, column=tax_col).value = "税额匹配"
        ws.cell(row=header_row, column=diff_col).value = "差额"
        max_row = ws.max_row
        end_row = max_row - 2
        for r in range(header_row + 1, end_row + 1):
            inv = str(ws.cell(r, target_col).value or "").strip()
            tax = b_map.get(inv, None)
            if tax:
                try:
                    ws.cell(r, tax_col).value = float(tax)
                except:
                    ws.cell(r, tax_col).value = tax

        comment = Comment("请检查：1、发票票种 2、发票是否已抵扣", "系统")
        max_row = ws.max_row
        end_row = max_row - 2
        for r in range(header_row + 1, end_row + 1):
            cell = ws.cell(r, tax_col)
            if cell.value in (None, ""):
                cell.fill = RED_CELL
                cell.comment = comment

        log("🔧 添加全表边框...", log_widget)
        for r in range(header_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).border = border

        log("🔧 删除 J、K、L 列...", log_widget)
        ws.delete_cols(10, 3)

        ws.cell(row=8, column=10).value = "发票税额"
        ws.cell(row=8, column=11).value = "差额"
        ws.cell(row=8, column=10).fill = WHITE_FILL
        if ws.cell(row=8, column=10).comment:
            ws.cell(row=8, column=10).comment = None
        ws.cell(row=7, column=10).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=7, column=11).alignment = Alignment(horizontal="center", vertical="center")

        log("🔧 最终合并差额单元格...", log_widget)
        voucher_groups = defaultdict(list)
        max_row = ws.max_row
        end_row = max_row - 2
        for r in range(9, end_row + 1):
            v = str(ws.cell(r, 3).value or "").strip()
            if v:
                voucher_groups[v].append(r)
        for voucher, rows in voucher_groups.items():
            s = min(rows)
            e = max(rows)
            sumI = 0.0
            sumJ = 0.0
            for r in range(s, e+1):
                try:
                    sumI += float(ws.cell(r,9).value or 0)
                except:
                    pass
                try:
                    sumJ += float(ws.cell(r,10).value or 0)
                except:
                    pass
            final_diff = round(sumI - sumJ, 2)
            ws.merge_cells(start_row=s, start_column=11, end_row=e, end_column=11)
            cell = ws.cell(s,11)
            cell.value = final_diff
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = "#,##0.00"
            if abs(final_diff) > 0.001:
                cell.fill = RED_CELL
                cell.comment = Comment("核对税额", "系统")

        log("🔧 统一格式 + 行高...", log_widget)
        tpl_row = 9
        sample = ws.cell(tpl_row, 1)
        std_height = ws.row_dimensions[tpl_row].height or 18
        std_font = Font(
            name=sample.font.name,
            size=sample.font.size,
            bold=sample.font.bold,
            italic=sample.font.italic,
            color=sample.font.color
        )
        max_row = ws.max_row
        end_row = max_row - 2
        for r in range(10, end_row + 1):
            ws.row_dimensions[r].height = std_height
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.fill in (RED, YELLOW, RED_CELL):
                    continue
                cell.font = std_font
                cell.number_format = sample.number_format
        ws.row_dimensions[tpl_row].height = std_height

        # 保存：系统「另存为」对话框
        default_name = "发票提取 + 税额核对完成表.xlsx"
        initial_dir = os.path.dirname(file_a) if file_a else os.getcwd()
        out = filedialog.asksaveasfilename(
            title="保存处理结果",
            initialdir=initial_dir,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not out:
            log("❌ 已取消保存，文件未写入磁盘", log_widget)
            return
        wb.save(out)
        log(f"\n🎉 全部处理完成！", log_widget)
        log(f"📁 已保存：{out}", log_widget)
        log("🔍 正在自动打开文件...", log_widget)
        os.startfile(out)
        messagebox.showinfo("完成", f"处理完成！\n文件已保存并打开：\n{out}")

    except Exception as e:
        err_msg = f"运行出错：{str(e)}\n{traceback.format_exc()}"
        log(f"❌ 出错了！详细信息：\n{err_msg}", log_widget)
        messagebox.showerror("错误", f"运行出错，程序不会退出，请查看日志：\n{str(e)}")
        for i in range(60, 0, -1):
            log(f"等待 {i} 秒后继续操作...", log_widget)
            time.sleep(1)
        log("已恢复，可继续操作", log_widget)

# ==================== 工具2：生成抵扣模板 ====================
def tool2_build_template(log_widget, file_a=None, file_b=None, file_c=None):
    try:
        log("==================================================", log_widget)
        log("    进项税抵扣勾选模板生成工具（财务专用版）", log_widget)
        log("    支持格式：.xlsx / .xls 全兼容", log_widget)
        log("==================================================", log_widget)
        log("", log_widget)

        # 选择文件
        if not file_a:
            log("📂 请选择【A表：原始账务表.xlsx】", log_widget)
            a = filedialog.askopenfilename(
                title="A表：原始账务表", 
                filetypes=[("Excel文件", "*.xlsx;*.xls")]
            )
        else:
            a = file_a
        if not a:
            log("❌ 未选择A表", log_widget)
            return
        
        if not file_b:
            log("📂 请选择【B表：抵扣勾选增值税发票信息.xlsx】", log_widget)
            b = filedialog.askopenfilename(
                title="B表：抵扣发票信息", 
                filetypes=[("Excel文件", "*.xlsx;*.xls")]
            )
        else:
            b = file_b
        if not b:
            log("❌ 未选择B表", log_widget)
            return
        
        if not file_c:
            log("📂 请选择【C表：进项税发票抵扣勾选导入模版.xlsx】", log_widget)
            c = filedialog.askopenfilename(
                title="C表：导入模板", 
                filetypes=[("Excel文件", "*.xlsx;*.xls")]
            )
        else:
            c = file_c
        if not c:
            log("❌ 未选择C表", log_widget)
            return

        log(f"✅ 已选择文件：\n  A表: {a}\n  B表: {b}\n  C表: {c}", log_widget)

        # 兼容加载
        wb_a = load_workbook_compat(a)
        ws_a = wb_a.active
        col = None
        for i in range(1,11):
            for j in range(1,30):
                if "摘要" in str(ws_a.cell(i,j).value or ""):
                    col = j
                    log(f"✅ 找到摘要列：第{i}行，第{j}列", log_widget)
                    break
            if col:
                break
        if not col:
            log("❌ 未找到摘要列", log_widget)
            messagebox.showerror("错误", "未找到摘要列")
            return

        log("🔧 提取发票号...", log_widget)
        invs = []
        end = ws_a.max_row - 2
        for r in range(9, end+1):
            s = str(ws_a.cell(r, col).value or "")
            for n in re.findall(r"\d+", s):
                if 15 <= len(n) <=30:
                    invs.append(n)
                    break

        if not invs:
            log("❌ 未提取到任何发票号", log_widget)
            messagebox.showwarning("提示", "未提取到发票")
            return
        log(f"✅ 共提取到 {len(invs)} 张发票", log_widget)

        log("🔧 加载B表数据...", log_widget)
        wb_b = load_workbook_compat(b)
        ws_b = wb_b.active
        mp = {}
        for r in range(2, ws_b.max_row+1):
            k = str(ws_b.cell(r,3).value or "").strip()
            if k:
                mp[k] = {
                    "数电发票号码": ws_b.cell(r,2).value,
                    "发票代码": ws_b.cell(r,4).value,
                    "开票日期*": ws_b.cell(r,6).value,
                    "金额*": ws_b.cell(r,7).value,
                    "票面税额*": ws_b.cell(r,8).value,
                    "有效抵扣税额*": ws_b.cell(r,9).value,
                    "购货方识别号*": ws_b.cell(r,10).value,
                    "销售方名称": ws_b.cell(r,11).value,
                    "销售方识别号*": ws_b.cell(r,12).value,
                    "发票来源": ws_b.cell(r,13).value,
                    "发票类型*": ws_b.cell(r,14).value,
                }
        log(f"✅ B表共加载 {len(mp)} 条发票信息", log_widget)

        log("🔧 填充模板数据...", log_widget)
        wb_c = load_workbook_compat(c)
        ws_c = wb_c.active
        cnt = Counter([str(x).strip() for x in invs])
        row = 2
        match_count = 0

        for v in invs:
            vs = str(v).strip()
            ws_c.cell(row,1,"是")
            cell = ws_c.cell(row,2,vs)

            if len(vs)!=20:
                cell.fill = RED_CELL
                cell.comment = Comment("错误","系统")
            if cnt[vs]>=2:
                cell.fill = YELLOW
                cell.comment = Comment("重复发票","系统")

            if vs in mp:
                d = mp[vs]
                ws_c.cell(row,3, d["发票代码"])
                ws_c.cell(row,5, d["开票日期*"])
                ws_c.cell(row,6, d["金额*"])
                ws_c.cell(row,7, d["票面税额*"])
                ws_c.cell(row,8, d["有效抵扣税额*"])
                ws_c.cell(row,9, d["购货方识别号*"])
                ws_c.cell(row,10, d["销售方名称"])
                ws_c.cell(row,11, d["销售方识别号*"])
                ws_c.cell(row,12, d["发票来源"])
                ws_c.cell(row,13, d["发票类型*"])
                match_count +=1

            for cc in range(1,14):
                ws_c.cell(row,cc).border = border
            row +=1

        log(f"✅ 模板填充完成：成功匹配 {match_count} 条", log_widget)

        # 保存：系统「另存为」对话框
        default_name = "✅已生成_进项税抵扣模板.xlsx"
        initial_dir = os.path.dirname(c) if c else os.getcwd()
        out = filedialog.asksaveasfilename(
            title="保存抵扣勾选模板",
            initialdir=initial_dir,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")],
        )
        if not out:
            log("❌ 已取消保存，文件未写入磁盘", log_widget)
            return
        wb_c.save(out)
        log("\n🎉 全部处理完成！", log_widget)
        log(f"📁 已保存：{out}", log_widget)
        log("🔍 正在自动打开文件...", log_widget)
        os.startfile(out)
        messagebox.showinfo("完成", f"处理完成！\n文件已保存并打开：\n{out}")

    except Exception as e:
        err_msg = f"运行出错：{str(e)}\n{traceback.format_exc()}"
        log(f"❌ 出错了！详细信息：\n{err_msg}", log_widget)
        messagebox.showerror("错误", f"运行出错，程序不会退出，请查看日志：\n{str(e)}")
        for i in range(60, 0, -1):
            log(f"等待 {i} 秒后继续操作...", log_widget)
            time.sleep(1)
        log("已恢复，可继续操作", log_widget)

# ==================== 分区域拖拽（A/B/C 各绑一个文件） ====================
def bind_zone_drop(widget, slot_key, path_state, text_label, log_widget):
    """每次拖入动作只更新对应分区；多文件时取第一个并提示。"""

    def on_drop(event):
        files = parse_drop_files(event.data)
        excel_files = [f for f in files if f.lower().endswith((".xlsx", ".xls"))]
        if not excel_files:
            messagebox.showwarning("提示", "只支持拖入 .xlsx / .xls 文件")
            return
        path = excel_files[0]
        if len(excel_files) > 1:
            log(
                f"⚠️ {slot_key.upper()} 区请每次拖入一个文件，已使用：{os.path.basename(path)}",
                log_widget,
            )
        path_state[slot_key] = path
        bn = os.path.basename(path)
        text_label.config(text=bn if len(bn) <= 40 else bn[:37] + "...")
        log(f"📥 {slot_key.upper()} 区：{path}", log_widget)

    widget.drop_target_register(DND_FILES)
    widget.dnd_bind("<<Drop>>", on_drop)


def make_drop_zone(parent, title, hint, slot_key, path_state, log_widget, width_chars=28):
    """创建可拖放的单文件区域，返回 Frame。"""
    fr = tk.Frame(
        parent,
        bg="#e8eef7",
        highlightbackground=BLUE,
        highlightthickness=2,
        bd=0,
    )
    tk.Label(fr, text=title, font=("微软雅黑", 11, "bold"), bg="#e8eef7", fg=BLUE).pack(
        pady=(10, 4)
    )
    name_lbl = tk.Label(
        fr,
        text="（拖入 Excel）",
        font=("微软雅黑", 9),
        bg="white",
        fg="#666666",
        width=width_chars,
        anchor="w",
        padx=8,
        pady=10,
    )
    name_lbl.pack(fill=tk.X, padx=10, pady=4)
    tk.Label(fr, text=hint, font=("微软雅黑", 8), bg="#e8eef7", fg="#888888").pack(
        pady=(0, 10)
    )
    bind_zone_drop(fr, slot_key, path_state, name_lbl, log_widget)
    return fr


# ==================== 主界面 ====================
def main():
    root = TkinterDnD.Tk()  # 支持拖拽的窗口
    root.title("发票抵扣工具集")
    root.geometry("800x820")
    root.resizable(False, False)
    root.configure(bg=BG_COLOR)

    path_state = {"a": None, "b": None, "c": None}

    # 先创建日志控件（拖放回调会写入日志），最后再 pack
    log_text = scrolledtext.ScrolledText(
        root,
        width=85,
        height=16,
        font=("Consolas", 10),
        bg="white",
        fg="#333333",
        bd=1,
        relief="solid",
    )
    log_text.config(state=tk.DISABLED)

    # 顶部标题
    tk.Label(
        root, text="📋 发票处理工具", font=("微软雅黑", 28, "bold"), fg=BLUE, bg=BG_COLOR
    ).pack(pady=16)

    tk.Label(
        root,
        text="💡 将文件分别拖到 A / B / C 区（每次一个）；再点下方对应按钮处理。也可不拖拽，直接点按钮选文件。",
        font=("微软雅黑", 9),
        fg="#666666",
        bg=BG_COLOR,
        wraplength=760,
        justify="center",
    ).pack(pady=(0, 8))

    # 拖放区：A | B
    row_ab = tk.Frame(root, bg=BG_COLOR)
    row_ab.pack(pady=8)
    zone_a = make_drop_zone(
        row_ab,
        "A 区",
        "原始账务表",
        "a",
        path_state,
        log_text,
        width_chars=32,
    )
    zone_a.pack(side=tk.LEFT, padx=10, ipadx=6, ipady=4)
    zone_b = make_drop_zone(
        row_ab,
        "B 区",
        "抵扣发票信息表",
        "b",
        path_state,
        log_text,
        width_chars=32,
    )
    zone_b.pack(side=tk.LEFT, padx=10, ipadx=6, ipady=4)

    # C 区（仅工具2需要；工具1忽略）
    zone_c = make_drop_zone(
        root,
        "C 区",
        "工具2：进项税抵扣勾选导入模版（仅功能2需要）",
        "c",
        path_state,
        log_text,
        width_chars=72,
    )
    zone_c.pack(pady=6, padx=20, ipadx=8, ipady=2, fill=tk.X)

    # 按钮区
    btn_frame = tk.Frame(root, bg=BG_COLOR)
    btn_frame.pack(pady=12)

    # 按钮1
    def on_btn1_enter(e):
        btn1.config(bg=BLUE, fg="white")

    def on_btn1_leave(e):
        btn1.config(bg="white", fg=BLUE)

    btn1 = tk.Button(
        btn_frame,
        text="📋  1. 发票提取 + 税额核对",
        font=("微软雅黑", 16),
        width=40,
        height=2,
        bg="white",
        fg=BLUE,
        bd=2,
        relief="flat",
        highlightbackground=BLUE,
        highlightthickness=2,
        command=lambda: tool1_extract_and_check(
            log_text, path_state["a"], path_state["b"]
        ),
    )
    btn1.pack(pady=10)
    btn1.bind("<Enter>", on_btn1_enter)
    btn1.bind("<Leave>", on_btn1_leave)

    # 按钮2
    def on_btn2_enter(e):
        btn2.config(bg=BLUE, fg="white")

    def on_btn2_leave(e):
        btn2.config(bg="white", fg=BLUE)

    btn2 = tk.Button(
        btn_frame,
        text="📋  2. 生成抵扣勾选模板",
        font=("微软雅黑", 16),
        width=40,
        height=2,
        bg="white",
        fg=BLUE,
        bd=2,
        relief="flat",
        highlightbackground=BLUE,
        highlightthickness=2,
        command=lambda: tool2_build_template(
            log_text, path_state["a"], path_state["b"], path_state["c"]
        ),
    )
    btn2.pack(pady=10)
    btn2.bind("<Enter>", on_btn2_enter)
    btn2.bind("<Leave>", on_btn2_leave)

    tk.Label(root, text="📜 运行日志", font=("微软雅黑", 12), fg=BLUE, bg=BG_COLOR).pack()
    log_text.pack(pady=10)
    log_text.config(state=tk.DISABLED)

    tk.Label(
        root,
        text="© 财务专用 | 全格式兼容 | 分区域拖拽 | 另存为后自动打开",
        font=("微软雅黑", 11),
        fg=BLUE,
        bg=BG_COLOR,
    ).pack(pady=5)

    log("欢迎使用发票处理工具集！", log_text)
    log("使用方法：", log_text)
    log("  👉 将 Excel 拖到 A/B（功能1）或 A/B/C（功能2），每次拖入一个文件。", log_text)
    log("  👉 点对应按钮开始处理；未拖入的区域会在处理时弹出文件选择框。", log_text)
    log("  👉 支持 .xlsx 与 .xls；处理结束后使用系统「另存为」选择路径与文件名，保存后自动打开。", log_text)
    log("-" * 60, log_text)

    root.mainloop()

if __name__ == "__main__":
    main()
